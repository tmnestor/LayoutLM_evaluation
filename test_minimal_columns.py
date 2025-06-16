from pathlib import Path

import openpyxl
import pandas as pd

# Test file path
file_path = Path('my_file.xlsx')

# Minimal column extraction approach
wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
ws = wb.active

# Find header row and required column indices
headers = []
pred_col = None
annotator1_col = None

for row in ws.iter_rows(max_row=1, values_only=True):
    headers = [str(cell) if cell is not None else f"col_{j}" for j, cell in enumerate(row)]
    break

for i, header in enumerate(headers):
    if header == 'pred':
        pred_col = i
    elif header == 'annotator1_label':
        annotator1_col = i

# Extract only the required columns
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    row_data = {}
    if pred_col is not None:
        row_data['pred'] = row[pred_col] if pred_col < len(row) else None
    if annotator1_col is not None:
        row_data['annotator1_label'] = row[annotator1_col] if annotator1_col < len(row) else None
    data.append(row_data)

wb.close()
test_df = pd.DataFrame(data)
print(f"Shape: {test_df.shape}")
print(test_df.head())